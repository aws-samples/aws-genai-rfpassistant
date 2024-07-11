import openpyxl
import boto3
import json
import io

class ExcelWorksheet:
    def __init__(self,sheet,sheet_name,starting_row,query_column):
        """The starting_row and query_column start from 0 for the first row/column"""
        self.starting_row = starting_row
        self.query_column = query_column
        self.sheet_name = sheet_name
        rows = list(sheet.values)
        self.queries_only = list()
        self.query_list = list()
        for every_row in rows[starting_row:]:
            if every_row[query_column] != '':
                self.queries_only.append(every_row[query_column])
            continue
        self.index = 0

    def get_all_queries(self):
        """Lets return a series of queries"""
        self.query_list = [dict(row=row_num+self.starting_row,query=q) for row_num,q in enumerate(self.queries_only,start=0)]
        return self.query_list

class ExcelSXReader:
    def __init__(self,stream_body,*,query_start_row=0,query_column=0):
        self.query_start_row = query_start_row ## Where the queries actually start from in every sheet in the workbook
        self.query_column = query_column
        self.wb = openpyxl.load_workbook(io.BytesIO(stream_body.read())) ## Skips empty columns and rows!
        self.sheets = [ExcelWorksheet(sheet, sheet.title, starting_row=self.query_start_row, query_column=self.query_column) for sheet in self.wb.worksheets]

    def get_number_of_sheets(self):
        return len(self.sheets)

    def _get_all_queries(self,sheet):
        all_queries = list()
        all_queries.append(sheet.get_all_queries())
        return all_queries

    def get_all_queries(self):
        list_of_queries = list()
        for sheet in self.sheets:
            all_queries = self._get_all_queries(sheet)
            ## create dict of sheet and every query in the all_queries
            list_of_queries.append([dict(sheet=sheet.sheet_name, Column=self.query_column, queries=query) for query in all_queries])
        return list_of_queries

class ExcelSXWriter:
    def __init__(self,stream_body,*,query_start_row=0,query_column=0):
        self.query_start_row = query_start_row 
        self.response_column = query_column + 1
        self.wb = openpyxl.load_workbook(io.BytesIO(stream_body.read()))

    def edit_response(self, workbook_object,destination_dir='/tmp/',batch=-1,start_from=0):
        """
        The workbook object is a JSON object
        This function is called internally to edit an excel
        Unless otherwise specific, the destination directory is /tmp/, considering this
        will be executed in an AWS Lambda function.
        Considering that, it may so happen that one 15 min run of a Lambda function may
        not be adequate to complete the task. We specify where we want to start from.
        The batch parameter is used to determine how many queries we want to process.
        If batch is -1, we process all the queries.
        We also return where we left off.
        """
        wb_obj = json.loads(workbook_object)
        for sheet in wb_obj:
            for some_sh in sheet:
                ws = self.wb[some_sh['sheet']]
                for query in some_sh['queries']:
                    if not query['feedbackresponse']:
                        ws.cell(row=query['row']+self.query_start_row, column=self.response_column+1).value = query['generatedresponse']
                        continue
                    ws.cell(row=query['row']+self.query_start_row, column=self.response_column+1).value = query['feedbackresponse']
        self.wb.save(destination_dir + 'output_file.xlsx')
        

## This is the driver
if __name__ == '__main__':
    file_path = r"Documents/Explorations/QnA_Samples/Queries_and_Responses.xlsx"
    s3 = boto3.client('s3',region_name='us-west-2')
    response = s3.get_object(Bucket='raneesai-demos-us-west-2024',Key='test/Queries_and_Responses.xlsx')
    reader = ExcelSXReader(response['Body'], query_start_row=3,query_column=2) ## query_start_row(starts from 1 if we asume headers) and query_column(starts from 0)
    #print(reader.get_all_queries())
    response = s3.get_object(Bucket='raneesai-demos-us-west-2024',Key='test/Queries_and_Responses.xlsx')
    writer = ExcelSXWriter(response['Body'], query_start_row=3, query_column=2)
    ## Load this JSON from file, for the sake of this test
    with open(r"Documents/Explorations/QnA_Samples/sample_response.json") as f:
        writer.edit_response(f,destination_dir=r"Documents/Explorations/QnA_Samples/")